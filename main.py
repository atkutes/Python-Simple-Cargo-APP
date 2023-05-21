import customtkinter
import openpyxl as xl
from tkinter import *
from tkinter import messagebox

#We need to import Tkinter
#App all settings
app = customtkinter.CTk()
app.title('Cargo Entry')
app.geometry('700x400')
app.config(bg='#000')

font1 = ('Arial', 25, 'bold')
font2 = ('Arial', 18, 'bold')

def clear():
    mobile_entry.delete(0,END)
    weight_entry.delete(0,END)
    variable1.set('')
    variable2.set('')

def submit():
    if mobile_entry.get() !='' and weight_entry.get() !='' and variable1.get() !='':
        file = xl.load_workbook('data.xlsx')
        sheet = file['Sheet1']
        mobile_value = mobile_entry.get()
        weight_value = weight_entry.get()
        from_to_value = variable1.get()
        delivery_value = variable2.get()
        sheet.cell(column=1,row=sheet.max_row+1,value=mobile_value)
        sheet.cell(column=2,row=sheet.max_row,value=weight_value)
        sheet.cell(column=3,row=sheet.max_row,value=from_to_value)
        sheet.cell(column=4,row=sheet.max_row,value=delivery_value)
        file.save('data.xlsx')
        messagebox.showinfo('Success', 'Data has been saved.')
    else:
        messagebox.showerror('Error', 'Enter all the data.')

#All buttons and Frames
frame = customtkinter.CTkFrame(app, fg_color='#110e21', bg_color='#000', width=650, height=350, border_color='#fff', border_width=3)
frame.place(x=10,y=20)

mobile_label = customtkinter.CTkLabel(frame, font=font1, text='Mobile Number:', text_color='#fff', bg_color='#110e21')
mobile_label.place(x=10,y=10)

mobile_entry = customtkinter.CTkEntry(frame, font=font2, text_color='#000', fg_color='#fff', border_color='#fff', width=150)
mobile_entry.place(x=200,y=10)

weight_label = customtkinter.CTkLabel(frame, font=font1, text='Weight:', text_color='#fff', bg_color='#110e21')
weight_label.place(x=10,y=60)

weight_entry = customtkinter.CTkEntry(frame, font=font2, text_color='#000', fg_color='#fff', border_color='#fff', width=150)
weight_entry.place(x=110,y=60)

from_to_label = customtkinter.CTkLabel(frame, font=font1, text='From/To:', text_color='#fff', bg_color='#110e21')
from_to_label.place(x=10,y=120)

from_to_options = ['A - B', 'B - A']
variable1 = StringVar()

from_to_option = customtkinter.CTkComboBox(frame, font=font2, text_color='#000', fg_color='#fff', dropdown_hover_color='#06911f', variable=variable1,values=from_to_options, state='readonly', width=150)
from_to_option.place(x=130,y=120)

delivery_label = customtkinter.CTkLabel(frame, font=font1, text='Delivery/Picking up:', text_color='#fff', bg_color='#110e21')
delivery_label.place(x=10,y=180)

variable2 = StringVar()

rb1 = customtkinter.CTkRadioButton(frame,text='Home', fg_color='#0000FF', text_color='#fff', hover_color='#0000FF', font=font2, variable=variable2, value='Home')
rb2 = customtkinter.CTkRadioButton(frame,text='Terminal', fg_color='#06c419',text_color='#fff', hover_color='#06c419', font=font2, variable=variable2, value='Terminal')

rb1.place(x=270,y=185)
rb2.place(x=380,y=185)


clear_button = customtkinter.CTkButton(app,command=clear, font=font1,text_color='#fff', text='Clear', fg_color='#d65104', hover_color='#913703',bg_color='#000', cursor='hand2',corner_radius=10,width=150)
clear_button.place(x=80,y=300)

submit_button = customtkinter.CTkButton(app, command=submit,font=font1,text_color='#fff', text='Submit', fg_color='#02ab10', hover_color='#913703',bg_color='#000',cursor='hand2',corner_radius=10,width=150)
submit_button.place(x=280,y=300)




app.mainloop()